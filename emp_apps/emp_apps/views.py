from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Q
from datetime import datetime, date, timedelta
from .models import Employee, AttendanceRecord, AttendanceSettings, LeaveRequest
from django.contrib.auth.models import User
from django.contrib.admin.views.decorators import staff_member_required
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def employee_login(request):
    """Employee login view"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Login successful!')
            # CUSTOMIZATION NOTE: After login, redirect to dashboard (attendance page)
            return redirect('attendance_dashboard')  # Changed from 'attendance_dashboard' to direct to dashboard
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'attendance/login.html')


def employee_logout(request):
    """Employee logout view"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('employee_login')


@login_required
def attendance_dashboard(request):
    """Main dashboard for employees to check in/out - This is the main attendance functionality"""
    try:
        employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found. Please contact administrator.')
        return redirect('employee_login')
    
    today = date.today()
    
    # Get today's attendance record
    today_record, created = AttendanceRecord.objects.get_or_create(
        employee=employee,
        date=today,
        defaults={'status': 'absent'}
    )
    
    # Get recent attendance records (last 7 days)
    recent_records = AttendanceRecord.objects.filter(
        employee=employee,
        date__gte=today - timedelta(days=7)
    ).order_by('-date')
    
    # Get attendance settings
    settings, _ = AttendanceSettings.objects.get_or_create(pk=1)
    
    context = {
        'employee': employee,
        'today_record': today_record,
        'recent_records': recent_records,
        'settings': settings,
        'current_time': timezone.now(),
    }
    
    return render(request, 'attendance/dashboard.html', context)


@login_required
def employee_profile(request):
    """Employee Profile view - Shows detailed employee information"""
    # CUSTOMIZATION NOTE: This view shows employee profile details
    # Admin can edit employee information through Django admin or you can add edit functionality here
    try:
        employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found. Please contact administrator.')
        return redirect('employee_login')
    
    # Get some basic statistics
    today = date.today()
    current_month_start = today.replace(day=1)
    
    # Current month statistics
    current_month_records = AttendanceRecord.objects.filter(
        employee=employee,
        date__gte=current_month_start,
        date__lte=today
    )
    
    present_days = current_month_records.filter(status__in=['present', 'late']).count()
    late_days = current_month_records.filter(status='late').count()
    total_work_hours = sum([r.work_hours or 0 for r in current_month_records])
    
    # Calculate years of service
    years_of_service = today.year - employee.hire_date.year
    # Adjust if birthday hasn't occurred this year
    if today.month < employee.hire_date.month or (today.month == employee.hire_date.month and today.day < employee.hire_date.day):
        years_of_service -= 1
    
    context = {
        'employee': employee,
        'user': request.user,
        'years_of_service': years_of_service,
        'statistics': {
            'present_days': present_days,
            'late_days': late_days,
            'total_work_hours': total_work_hours,
            'month_name': current_month_start.strftime('%B %Y')
        }
    }
    
    return render(request, 'attendance/employee_profile.html', context)


@login_required
def check_in(request):
    """Handle employee check-in"""
    if request.method == 'POST':
        try:
            employee = Employee.objects.get(user=request.user)
            today = date.today()
            current_time = timezone.now()
            
            # Get attendance settings
            settings, _ = AttendanceSettings.objects.get_or_create(pk=1)
            
            # Check if check-in is allowed at current time
            if not settings.is_check_in_allowed(current_time):
                return JsonResponse({
                    'success': False,
                    'message': f'Check-in not allowed at this time. {settings.get_check_in_window_message()}'
                })
            
            # Get or create today's attendance record
            attendance_record, created = AttendanceRecord.objects.get_or_create(
                employee=employee,
                date=today,
                defaults={'status': 'absent'}
            )
            
            # Check if already checked in
            if attendance_record.check_in_time:
                return JsonResponse({
                    'success': False,
                    'message': 'You have already checked in today.'
                })
            
            # Set check-in time
            attendance_record.check_in_time = current_time
            
            # Determine if late
            if attendance_record.is_late(settings.standard_check_in_time.strftime('%H:%M')):
                attendance_record.status = 'late'
                status_message = f'Checked in at {current_time.strftime("%H:%M:%S")} (Late)'
            else:
                attendance_record.status = 'present'
                status_message = f'Checked in successfully at {current_time.strftime("%H:%M:%S")}'
            
            attendance_record.save()
            
            return JsonResponse({
                'success': True,
                'message': status_message,
                'check_in_time': attendance_record.check_in_time.strftime('%H:%M:%S'),
                'status': attendance_record.status
            })
            
        except Employee.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Employee profile not found.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


@login_required
def check_out(request):
    """Handle employee check-out"""
    if request.method == 'POST':
        try:
            employee = Employee.objects.get(user=request.user)
            today = date.today()
            current_time = timezone.now()
            
            # Get attendance settings
            settings, _ = AttendanceSettings.objects.get_or_create(pk=1)
            
            # Check if check-out is allowed at current time
            if not settings.is_check_out_allowed(current_time):
                return JsonResponse({
                    'success': False,
                    'message': f'Check-out not allowed at this time. {settings.get_check_out_window_message()}'
                })
            
            # Get today's attendance record
            try:
                attendance_record = AttendanceRecord.objects.get(
                    employee=employee,
                    date=today
                )
            except AttendanceRecord.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'No check-in record found for today.'
                })
            
            # Check if not checked in
            if not attendance_record.check_in_time:
                return JsonResponse({
                    'success': False,
                    'message': 'You need to check in first.'
                })
            
            # Check if already checked out
            if attendance_record.check_out_time:
                return JsonResponse({
                    'success': False,
                    'message': 'You have already checked out today.'
                })
            
            # Check minimum work time (optional - prevent immediate check-out)
            time_since_checkin = current_time - attendance_record.check_in_time
            minimum_work_minutes = 30  # Minimum 30 minutes between check-in and check-out
            if time_since_checkin.total_seconds() < (minimum_work_minutes * 60):
                return JsonResponse({
                    'success': False,
                    'message': f'You must work at least {minimum_work_minutes} minutes before checking out.'
                })
            
            # Set check-out time and calculate work hours
            attendance_record.check_out_time = current_time
            work_hours = attendance_record.calculate_work_hours()
            attendance_record.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Checked out successfully at {current_time.strftime("%H:%M:%S")}',
                'check_out_time': attendance_record.check_out_time.strftime('%H:%M:%S'),
                'work_hours': str(work_hours)
            })
            
        except Employee.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Employee profile not found.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


@login_required
def attendance_history(request):
    """View attendance history for the logged-in employee"""
    try:
        employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('employee_login')
    
    # Get date range from request (default to current month)
    today = date.today()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = today.replace(day=1)  # First day of current month
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = today
    
    # Get attendance records for the date range
    attendance_records = AttendanceRecord.objects.filter(
        employee=employee,
        date__range=[start_date, end_date]
    ).order_by('-date')
    
    # Calculate statistics
    total_days = (end_date - start_date).days + 1
    present_days = attendance_records.filter(status__in=['present', 'late']).count()
    late_days = attendance_records.filter(status='late').count()
    absent_days = attendance_records.filter(status='absent').count()
    total_work_hours = sum([r.work_hours or 0 for r in attendance_records])
    
    context = {
        'employee': employee,
        'attendance_records': attendance_records,
        'start_date': start_date,
        'end_date': end_date,
        'statistics': {
            'total_days': total_days,
            'present_days': present_days,
            'late_days': late_days,
            'absent_days': absent_days,
            'total_work_hours': total_work_hours,
        }
    }
    
    return render(request, 'attendance/history.html', context)


@login_required
def get_attendance_status(request):
    """AJAX endpoint to get current attendance status"""
    try:
        employee = Employee.objects.get(user=request.user)
        today = date.today()
        current_time = timezone.now()
        
        # Get attendance settings
        settings, _ = AttendanceSettings.objects.get_or_create(pk=1)
        
        try:
            attendance_record = AttendanceRecord.objects.get(
                employee=employee,
                date=today
            )
            
            # Check if actions are allowed at current time
            can_check_in = settings.is_check_in_allowed(current_time) and not attendance_record.check_in_time
            can_check_out = settings.is_check_out_allowed(current_time) and attendance_record.check_in_time and not attendance_record.check_out_time
            
            return JsonResponse({
                'success': True,
                'checked_in': bool(attendance_record.check_in_time),
                'checked_out': bool(attendance_record.check_out_time),
                'check_in_time': attendance_record.check_in_time.strftime('%H:%M:%S') if attendance_record.check_in_time else None,
                'check_out_time': attendance_record.check_out_time.strftime('%H:%M:%S') if attendance_record.check_out_time else None,
                'status': attendance_record.status,
                'work_hours': str(attendance_record.work_hours) if attendance_record.work_hours else '0.00',
                'can_check_in': can_check_in,
                'can_check_out': can_check_out,
                'check_in_window': settings.get_check_in_window_message(),
                'check_out_window': settings.get_check_out_window_message(),
                'current_time': current_time.strftime('%H:%M:%S')
            })
            
        except AttendanceRecord.DoesNotExist:
            # Check if actions are allowed at current time
            can_check_in = settings.is_check_in_allowed(current_time)
            can_check_out = False
            
            return JsonResponse({
                'success': True,
                'checked_in': False,
                'checked_out': False,
                'check_in_time': None,
                'check_out_time': None,
                'status': 'absent',
                'work_hours': '0.00',
                'can_check_in': can_check_in,
                'can_check_out': can_check_out,
                'check_in_window': settings.get_check_in_window_message(),
                'check_out_window': settings.get_check_out_window_message(),
                'current_time': current_time.strftime('%H:%M:%S')
            })
            
    except Employee.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Employee profile not found.'
        })


# Employee Leave Request views
@login_required
def submit_leave_request(request):
    """Submit a leave request"""
    if request.method == 'POST':
        try:
            employee = Employee.objects.get(user=request.user)
            data = json.loads(request.body)
            
            # Create leave request
            leave_request = LeaveRequest.objects.create(
                employee=employee,
                leave_type=data.get('leave_type'),
                start_date=datetime.strptime(data.get('start_date'), '%Y-%m-%d').date(),
                end_date=datetime.strptime(data.get('end_date'), '%Y-%m-%d').date(),
                reason=data.get('reason'),
                status='pending'
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Leave request submitted successfully! You will be notified once it is reviewed.'
            })
            
        except Employee.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Employee profile not found.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error submitting leave request: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method.'
    })


@login_required
def employee_leave_requests(request):
    """View employee's leave requests"""
    try:
        employee = Employee.objects.get(user=request.user)
        leave_requests = LeaveRequest.objects.filter(employee=employee).order_by('-applied_date')
        
        context = {
            'employee': employee,
            'leave_requests': leave_requests,
        }
        
        return render(request, 'attendance/leave_requests.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('employee_login')


# Admin views
@staff_member_required
def admin_dashboard(request):
    """Admin dashboard with employee management, attendance reports, and leave requests"""
    
    # Get statistics
    total_employees = Employee.objects.filter(is_active=True).count()
    today = date.today()
    present_today = AttendanceRecord.objects.filter(
        date=today, 
        status__in=['present', 'late']
    ).count()
    pending_leaves = LeaveRequest.objects.filter(status='pending').count()
    active_leaves = LeaveRequest.objects.filter(
        status='approved',
        start_date__lte=today,
        end_date__gte=today
    ).count()
    
    # Get all employees
    employees = Employee.objects.filter(is_active=True).select_related('user').order_by('user__first_name')
    
    # Get recent leave requests
    leave_requests = LeaveRequest.objects.select_related('employee__user').order_by('-applied_date')[:50]
    
    # Date range for attendance (default last 30 days)
    default_end_date = today
    default_start_date = today - timedelta(days=30)
    
    context = {
        'total_employees': total_employees,
        'present_today': present_today,
        'pending_leaves': pending_leaves,
        'active_leaves': active_leaves,
        'employees': employees,
        'leave_requests': leave_requests,
        'default_start_date': default_start_date.strftime('%Y-%m-%d'),
        'default_end_date': default_end_date.strftime('%Y-%m-%d'),
    }
    
    return render(request, 'attendance/admin_dashboard.html', context)


@staff_member_required
def admin_get_attendance(request):
    """AJAX endpoint to get attendance data with filters"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            employee_id = data.get('employee_id')
            
            # Build query
            query = Q()
            if start_date:
                query &= Q(date__gte=start_date)
            if end_date:
                query &= Q(date__lte=end_date)
            if employee_id:
                query &= Q(employee_id=employee_id)
            
            attendance_records = AttendanceRecord.objects.filter(query).select_related(
                'employee__user'
            ).order_by('-date', 'employee__user__first_name')
            
            records_data = []
            for record in attendance_records:
                records_data.append({
                    'employee_name': record.employee.user.get_full_name() or record.employee.user.username,
                    'date': record.date.strftime('%b %d, %Y'),
                    'check_in_time': record.check_in_time.strftime('%H:%M:%S') if record.check_in_time else None,
                    'check_out_time': record.check_out_time.strftime('%H:%M:%S') if record.check_out_time else None,
                    'work_hours': str(record.work_hours) if record.work_hours else '0.00',
                    'status': record.status,
                    'status_display': record.get_status_display(),
                    'notes': record.notes,
                })
            
            return JsonResponse({
                'success': True,
                'attendance_records': records_data
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error loading attendance data: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@staff_member_required
def admin_add_employee(request):
    """AJAX endpoint to add new employee"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Check if username already exists
            if User.objects.filter(username=data['username']).exists():
                return JsonResponse({
                    'success': False,
                    'message': 'Username already exists'
                })
            
            # Check if employee ID already exists
            if Employee.objects.filter(employee_id=data['employee_id']).exists():
                return JsonResponse({
                    'success': False,
                    'message': 'Employee ID already exists'
                })
            
            # Create user
            user = User.objects.create_user(
                username=data['username'],
                email=data['email'],
                password=data['password'],
                first_name=data['first_name'],
                last_name=data['last_name']
            )
            
            # Create employee
            employee = Employee.objects.create(
                user=user,
                employee_id=data['employee_id'],
                department=data['department'],
                position=data['position'],
                phone_number=data.get('phone_number', ''),
                hire_date=data['hire_date']
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Employee {user.get_full_name()} added successfully!'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error adding employee: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@staff_member_required
def admin_update_leave_status(request):
    """AJAX endpoint to update leave request status"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            leave_id = data.get('leave_id')
            status = data.get('status')
            comments = data.get('comments', '')
            
            leave_request = get_object_or_404(LeaveRequest, id=leave_id)
            leave_request.status = status
            leave_request.approved_by = request.user
            leave_request.approved_date = timezone.now()
            leave_request.admin_comments = comments
            leave_request.save()
            
            action = 'approved' if status == 'approved' else 'rejected'
            return JsonResponse({
                'success': True,
                'message': f'Leave request {action} successfully!'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error updating leave status: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@staff_member_required
def admin_get_leave_details(request, leave_id):
    """AJAX endpoint to get leave request details"""
    try:
        leave_request = get_object_or_404(LeaveRequest, id=leave_id)
        
        details_html = f"""
        <div class="row">
            <div class="col-md-6">
                <strong>Employee:</strong> {leave_request.employee.user.get_full_name()}<br>
                <strong>Employee ID:</strong> {leave_request.employee.employee_id}<br>
                <strong>Department:</strong> {leave_request.employee.department}<br>
                <strong>Leave Type:</strong> {leave_request.get_leave_type_display()}<br>
            </div>
            <div class="col-md-6">
                <strong>Start Date:</strong> {leave_request.start_date.strftime('%B %d, %Y')}<br>
                <strong>End Date:</strong> {leave_request.end_date.strftime('%B %d, %Y')}<br>
                <strong>Duration:</strong> {leave_request.duration_days} day{"s" if leave_request.duration_days > 1 else ""}<br>
                <strong>Applied Date:</strong> {leave_request.applied_date.strftime('%B %d, %Y')}<br>
            </div>
        </div>
        <div class="row mt-3">
            <div class="col-12">
                <strong>Reason:</strong><br>
                <p class="mt-1">{leave_request.reason}</p>
            </div>
        </div>
        """
        
        if leave_request.admin_comments:
            details_html += f"""
            <div class="row mt-3">
                <div class="col-12">
                    <strong>Admin Comments:</strong><br>
                    <p class="mt-1">{leave_request.admin_comments}</p>
                </div>
            </div>
            """
        
        actions_html = ""
        if leave_request.status == 'pending':
            actions_html = f"""
            <button type="button" class="btn btn-success" onclick="approveLeave({leave_request.id})">
                <i class="fas fa-check me-2"></i>Approve
            </button>
            <button type="button" class="btn btn-danger" onclick="rejectLeave({leave_request.id})">
                <i class="fas fa-times me-2"></i>Reject
            </button>
            """
        
        actions_html += '<button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Close</button>'
        
        return JsonResponse({
            'success': True,
            'html': details_html,
            'actions_html': actions_html
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading leave details: {str(e)}'
        })


@staff_member_required
def export_attendance_excel(request):
    """Export attendance data to Excel file"""
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        employee_id = request.GET.get('employee_id')
        
        # Build query
        query = Q()
        if start_date:
            query &= Q(date__gte=start_date)
        if end_date:
            query &= Q(date__lte=end_date)
        if employee_id:
            query &= Q(employee_id=employee_id)
        
        attendance_records = AttendanceRecord.objects.filter(query).select_related(
            'employee__user'
        ).order_by('-date', 'employee__user__first_name')
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Header row
        headers = ['Employee Name', 'Employee ID', 'Department', 'Date', 'Check In', 'Check Out', 'Work Hours', 'Status', 'Notes']
        ws.append(headers)
        
        # Style header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for record in attendance_records:
            row_data = [
                record.employee.user.get_full_name() or record.employee.user.username,
                record.employee.employee_id,
                record.employee.department,
                record.date.strftime('%Y-%m-%d'),
                record.check_in_time.strftime('%H:%M:%S') if record.check_in_time else '',
                record.check_out_time.strftime('%H:%M:%S') if record.check_out_time else '',
                str(record.work_hours) if record.work_hours else '0.00',
                record.get_status_display(),
                record.notes
            ]
            ws.append(row_data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"attendance_report_{start_date or 'all'}_{end_date or 'all'}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting attendance data: {str(e)}')
        return redirect('admin_dashboard')